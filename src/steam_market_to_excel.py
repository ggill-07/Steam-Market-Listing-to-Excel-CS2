#!/usr/bin/env python3
"""Export Steam Community Market CS2 listings to an Excel file.

This script crawls all listing pages (up to 100 listings per page) for a given CS2 market
hash name, resolves each listing's inspect link, extracts metadata directly from
Steam's asset payload, and writes an Excel sheet with:
- float and wear
- paint seed
- page number
- price
- sticker presence
"""

from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import quote

import pandas as pd
import requests
from openpyxl.utils import get_column_letter


def get_runtime_project_dir() -> Path:
    """Resolve the project directory for both source runs and bundled desktop builds."""

    if getattr(sys, "frozen", False):
        executable_dir = Path(sys.executable).resolve().parent
        # When the bundled desktop app lives in "<project>\\dist", keep exports in
        # the project root instead of nesting them under dist/.
        if executable_dir.name.lower() == "dist" and (executable_dir.parent / "src").exists():
            return executable_dir.parent
        return executable_dir
    return Path(__file__).resolve().parents[1]


STEAM_APP_ID = 730
STEAM_CONTEXT_ID = 2
PAGE_SIZE = 100
DEFAULT_STEAM_PAGE_DELAY = 0.5
DEFAULT_STEAM_RETRIES = 5
PROPID_PATTERN = re.compile(r"%propid:(\d+)%")
ITEM_NAMEID_PATTERN = re.compile(r"Market_LoadOrderSpread\(\s*(\d+)\s*\)")
RETRIABLE_STATUS_CODES = {429, 500, 502, 503, 504}
DEFAULT_OUTPUT_DIR = get_runtime_project_dir() / "exports"
SKIN_EXPORT_SUBDIR = "skins"
CASE_EXPORT_SUBDIR = "cases"
STICKER_EXPORT_SUBDIR = "stickers"
LATEST_POINTER_FILENAME = ".latest_export.txt"
SUPPORTED_TABLE_SUFFIXES = {".csv", ".xlsx", ".xls"}
CLI_COMMANDS = {"fetch", "fetch-many", "sort", "filter", "stats", "show", "use"}
DEFAULT_FETCH_MANY_WORKERS = 3
RIGHT_ALIGN_COLUMNS = {"#", "page", "float", "price", "paint_seed", "sticker_count"}
STEAM_SESSION_REFRESH_PAGE_INTERVAL = 10
STEAM_RECOVERY_WAIT_STEPS = (30.0, 60.0, 120.0)
STEAM_RECOVERY_PAGE_DELAY_FLOOR = 1.0
STATTRAK_PREFIX_PATTERN = re.compile(
    r"^\s*(?:stattrak|stattrack)(?:\u2122)?\s+",
    re.IGNORECASE,
)
DEFAULT_SHOW_COLUMNS = [
    "page",
    "float",
    "price",
    "stickers",
    "wear",
    "paint_seed",
    "listing_id",
]
WEAR_OPTIONS = (
    "Factory New",
    "Minimal Wear",
    "Field-Tested",
    "Well-Worn",
    "Battle-Scarred",
)


@dataclass
class ListingRow:
    listing_id: str
    asset_id: str
    page: int
    price: float
    currency: str
    float_value: Optional[float]
    wear: Optional[str]
    paint_seed: Optional[int]
    has_stickers: Optional[bool]
    sticker_count: Optional[int]
    inspect_link: Optional[str]


@dataclass
class FetchResult:
    market_hash_name: str
    output_path: Path
    dataframe: pd.DataFrame
    change_summary: Optional[Dict[str, int]]
    summary_override: Optional[str] = None


def get_wear_from_float(float_value: Optional[float]) -> Optional[str]:
    if float_value is None:
        return None
    if float_value < 0.07:
        return "Factory New"
    if float_value < 0.15:
        return "Minimal Wear"
    if float_value < 0.38:
        return "Field-Tested"
    if float_value < 0.45:
        return "Well-Worn"
    return "Battle-Scarred"


def normalize_market_hash_name_input(market_hash_name: str) -> str:
    normalized_name = market_hash_name.strip()
    if STATTRAK_PREFIX_PATTERN.match(normalized_name):
        normalized_name = STATTRAK_PREFIX_PATTERN.sub(
            "StatTrak\u2122 ",
            normalized_name,
            count=1,
        )
    return normalized_name


def extract_wear_name_from_market_hash_name(market_hash_name: str) -> Optional[str]:
    normalized_name = normalize_market_hash_name_input(market_hash_name)
    for wear_name in WEAR_OPTIONS:
        if normalized_name.endswith(f" ({wear_name})"):
            return wear_name
    return None


def market_item_supports_wear(market_hash_name: str) -> bool:
    return extract_wear_name_from_market_hash_name(market_hash_name) is not None


def coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def coerce_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_price_text(price_text: Any) -> Optional[float]:
    if not isinstance(price_text, str):
        return None
    match = re.search(r"(\d[\d,]*)(?:\.(\d{1,2}))?", price_text)
    if not match:
        return None
    whole_part = match.group(1).replace(",", "")
    decimal_part = match.group(2) or "0"
    normalized_text = f"{whole_part}.{decimal_part}"
    try:
        return float(normalized_text)
    except ValueError:
        return None


def extract_market_level_lowest_price(payload: Dict[str, Any]) -> Optional[float]:
    for key in ("lowest_price", "lowest_price_text", "sell_price_text"):
        parsed_price = parse_price_text(payload.get(key))
        if parsed_price is not None and parsed_price > 0:
            return parsed_price
    sell_price = coerce_float(payload.get("sell_price"))
    if sell_price is not None and sell_price > 0:
        return sell_price / 100.0
    return None


def extract_lowest_histogram_price(payload: Dict[str, Any]) -> Optional[float]:
    lowest_sell_order = coerce_float(payload.get("lowest_sell_order"))
    if lowest_sell_order is not None and lowest_sell_order > 0:
        return lowest_sell_order / 100.0

    sell_order_graph = payload.get("sell_order_graph") or []
    if sell_order_graph:
        first_row = sell_order_graph[0]
        if isinstance(first_row, list) and first_row:
            return coerce_float(first_row[0])

    return None


def extract_listing_total_price(
    listing: Dict[str, Any],
    market_level_fallback_price: Optional[float] = None,
) -> Optional[float]:
    price_cents = (listing.get("converted_price") or listing.get("price") or 0) + (
        listing.get("converted_fee") or listing.get("fee") or 0
    )
    if price_cents:
        price = float(price_cents) / 100.0
        if price > 0:
            return price

    if market_level_fallback_price is not None and market_level_fallback_price > 0:
        return market_level_fallback_price

    return None


def get_asset_property_lookup(asset_payload: Dict[str, Any]) -> Dict[int, Dict[str, Any]]:
    lookup: Dict[int, Dict[str, Any]] = {}
    for prop in asset_payload.get("asset_properties", []) or []:
        property_id = coerce_int(prop.get("propertyid"))
        if property_id is not None:
            lookup[property_id] = prop
    return lookup


def normalize_inspect_link(
    raw_link: str,
    listing_id: str,
    asset_id: str,
    asset_payload: Optional[Dict[str, Any]] = None,
) -> str:
    normalized = raw_link.replace("%listingid%", listing_id).replace("%assetid%", asset_id)
    property_lookup = get_asset_property_lookup(asset_payload or {})

    def replace_propid(match: re.Match[str]) -> str:
        property_id = int(match.group(1))
        prop = property_lookup.get(property_id, {})
        for key in ("string_value", "int_value", "float_value"):
            value = prop.get(key)
            if value is not None:
                return str(value)
        return match.group(0)

    return PROPID_PATTERN.sub(replace_propid, normalized)


def extract_steam_metadata(asset_payload: Dict[str, Any]) -> Dict[str, Any]:
    property_lookup = get_asset_property_lookup(asset_payload)

    float_value = coerce_float(property_lookup.get(2, {}).get("float_value"))
    paint_seed = coerce_int(property_lookup.get(1, {}).get("int_value"))

    descriptions = asset_payload.get("descriptions", []) or []
    sticker_count = sum(
        1
        for description in descriptions
        if isinstance(description, dict)
        and isinstance(description.get("value"), str)
        and "Sticker:" in description["value"]
    )
    if sticker_count:
        has_stickers: Optional[bool] = True
    else:
        has_stickers = None
        sticker_count = None

    return {
        "float_value": float_value,
        "paint_seed": paint_seed,
        "has_stickers": has_stickers,
        "sticker_count": sticker_count,
    }


def steam_render_page(
    session: requests.Session,
    market_hash_name: str,
    start: int,
    currency: int,
    country: str,
    language: str,
    max_retries: int = DEFAULT_STEAM_RETRIES,
) -> Dict[str, Any]:
    market_hash_name = normalize_market_hash_name_input(market_hash_name)
    encoded_name = quote(market_hash_name, safe="")
    url = f"https://steamcommunity.com/market/listings/{STEAM_APP_ID}/{encoded_name}/render/"
    params = {
        "start": start,
        "count": PAGE_SIZE,
        "currency": currency,
        "language": language,
        "country": country,
        "format": "json",
    }
    attempt = 0

    while True:
        response = session.get(url, params=params, timeout=25)
        if response.status_code not in RETRIABLE_STATUS_CODES:
            response.raise_for_status()
            payload = response.json()
            if not payload.get("success", False):
                # Steam sometimes answers late pages with success=false once the
                # market count has shifted underneath us. For non-first pages,
                # treat an empty payload as "we reached the end" instead of a
                # hard failure.
                if (
                    start > 0
                    and int(payload.get("total_count", 0) or 0) == 0
                    and not payload.get("listinginfo")
                ):
                    return payload
                raise RuntimeError(
                    f"Steam render endpoint returned unsuccessful response for start={start}"
                )
            return payload

        attempt += 1
        if attempt > max_retries:
            raise requests.HTTPError(
                f"Steam returned repeated temporary errors after {max_retries} retries for start={start}",
                response=response,
            )

        retry_after = response.headers.get("Retry-After")
        try:
            wait_seconds = float(
                retry_after) if retry_after is not None else 0.0
        except ValueError:
            wait_seconds = 0.0

        wait_seconds = max(wait_seconds, min(5 * attempt, 30))
        status = response.status_code
        print(
            f"Steam returned HTTP {status} for page starting at {start}. "
            f"Waiting {wait_seconds:.1f}s before retry {attempt}/{max_retries}..."
        )
        time.sleep(wait_seconds)


def extract_inspect_link(asset_payload: Dict[str, Any], listing_id: str, asset_id: str) -> Optional[str]:
    for key in ("market_actions", "actions"):
        actions = asset_payload.get(key) or []
        for action in actions:
            link = action.get("link")
            if isinstance(link, str) and ("csgo_econ_action_preview" in link or "%assetid%" in link):
                return normalize_inspect_link(
                    link,
                    listing_id=listing_id,
                    asset_id=asset_id,
                    asset_payload=asset_payload,
                )
    return None


def iter_listings(
    session: requests.Session,
    market_hash_name: str,
    currency: int,
    country: str,
    language: str,
    steam_page_delay: float = DEFAULT_STEAM_PAGE_DELAY,
    steam_max_retries: int = DEFAULT_STEAM_RETRIES,
) -> Iterable[ListingRow]:
    start = 0
    total_count: Optional[int] = None
    effective_page_delay = steam_page_delay
    pages_fetched = 0

    try:
        while total_count is None or start < total_count:
            if should_refresh_steam_session(pages_fetched):
                print(
                    f"Refreshing the Steam session after {pages_fetched} listing pages "
                    f"to reduce late-run rate limiting..."
                )
                close_requests_session(session)
                session = create_requests_session()

            try:
                payload = steam_render_page(
                    session=session,
                    market_hash_name=market_hash_name,
                    start=start,
                    currency=currency,
                    country=country,
                    language=language,
                    max_retries=steam_max_retries,
                )
            except requests.HTTPError as exc:
                response = getattr(exc, "response", None)
                if getattr(response, "status_code", None) not in RETRIABLE_STATUS_CODES:
                    raise
                close_requests_session(session)
                payload, session, effective_page_delay = recover_steam_render_page(
                    session=session,
                    market_hash_name=market_hash_name,
                    start=start,
                    currency=currency,
                    country=country,
                    language=language,
                    max_retries=steam_max_retries,
                    steam_page_delay=effective_page_delay,
                )

            total_count = int(payload.get("total_count", 0))
            listing_info = payload.get("listinginfo", {})
            if not listing_info:
                break

            assets = payload.get("assets", {}).get(
                str(STEAM_APP_ID), {}).get(str(STEAM_CONTEXT_ID), {})

            page_number = (start // PAGE_SIZE) + 1
            for listing_id, listing in listing_info.items():
                asset = listing.get("asset") or {}
                asset_id = str(asset.get("id", ""))
                asset_payload = assets.get(asset_id, {})

                inspect_link = extract_inspect_link(
                    asset_payload, listing_id=listing_id, asset_id=asset_id)

                price = extract_listing_total_price(listing)
                if price is None:
                    continue

                steam_metadata = extract_steam_metadata(asset_payload)
                float_value = steam_metadata["float_value"]
                paint_seed = steam_metadata["paint_seed"]
                has_stickers = steam_metadata["has_stickers"]
                sticker_count = steam_metadata["sticker_count"]

                yield ListingRow(
                    listing_id=listing_id,
                    asset_id=asset_id,
                    page=page_number,
                    price=price,
                    currency=str(listing.get("currencyid", currency)),
                    float_value=float_value,
                    wear=get_wear_from_float(float_value),
                    paint_seed=paint_seed,
                    has_stickers=has_stickers,
                    sticker_count=sticker_count,
                    inspect_link=inspect_link,
                )

            pages_fetched += 1
            start += PAGE_SIZE
            # Be nice to Steam and avoid hammering listing pages.
            time.sleep(effective_page_delay)
    finally:
        close_requests_session(session)


def rows_to_dataframe(rows: List[ListingRow]) -> pd.DataFrame:
    records = []
    for row in rows:
        records.append(
            {
                "listing_id": row.listing_id,
                "asset_id": row.asset_id,
                "page": row.page,
                "price": row.price,
                "currency": row.currency,
                "float": row.float_value,
                "wear": row.wear,
                "paint_seed": row.paint_seed,
                "has_stickers": row.has_stickers,
                "sticker_count": row.sticker_count,
                "inspect_link": row.inspect_link,
            }
        )
    return pd.DataFrame.from_records(records)


def attach_fetch_timestamp_columns(
    dataframe: pd.DataFrame,
    timestamp: Optional[datetime] = None,
) -> pd.DataFrame:
    attached_dataframe = dataframe.copy()
    timestamp = timestamp or datetime.now().astimezone()
    attached_dataframe["snapshot_date"] = timestamp.date().isoformat()
    attached_dataframe["snapshot_timestamp"] = timestamp.isoformat(timespec="seconds")
    return attached_dataframe


def build_lowest_listing_snapshot_dataframe(
    session: requests.Session,
    market_hash_name: str,
    currency: int,
    country: str,
    language: str,
    steam_max_retries: int = DEFAULT_STEAM_RETRIES,
) -> pd.DataFrame:
    commodity_lowest_price, commodity_price_source = fetch_commodity_lowest_price(
        session=session,
        market_hash_name=market_hash_name,
        currency=currency,
        country=country,
        language=language,
    )
    payload = steam_render_page(
        session=session,
        market_hash_name=market_hash_name,
        start=0,
        currency=currency,
        country=country,
        language=language,
        max_retries=steam_max_retries,
    )
    listing_info = payload.get("listinginfo", {})
    market_level_lowest_price = extract_market_level_lowest_price(payload)

    if not listing_info:
        if commodity_lowest_price is None:
            return attach_fetch_timestamp_columns(rows_to_dataframe([]))

        snapshot_dataframe = rows_to_dataframe(
            [
                ListingRow(
                    listing_id="",
                    asset_id="",
                    page=1,
                    price=commodity_lowest_price,
                    currency=str(currency),
                    float_value=None,
                    wear=None,
                    paint_seed=None,
                    has_stickers=None,
                    sticker_count=None,
                    inspect_link=None,
                )
            ]
        )
        snapshot_dataframe["price_source"] = commodity_price_source or "commodity_fallback"
        return attach_fetch_timestamp_columns(snapshot_dataframe)

    assets = payload.get("assets", {}).get(str(STEAM_APP_ID), {}).get(str(STEAM_CONTEXT_ID), {})
    priced_listings: List[tuple[str, Dict[str, Any], float]] = []
    for listing_id, listing in listing_info.items():
        listing_price = extract_listing_total_price(
            listing,
            market_level_fallback_price=market_level_lowest_price,
        )
        if listing_price is None:
            continue
        priced_listings.append((str(listing_id), listing, listing_price))

    if not priced_listings:
        if commodity_lowest_price is None:
            return attach_fetch_timestamp_columns(rows_to_dataframe([]))
        snapshot_dataframe = rows_to_dataframe(
            [
                ListingRow(
                    listing_id="",
                    asset_id="",
                    page=1,
                    price=commodity_lowest_price,
                    currency=str(currency),
                    float_value=None,
                    wear=None,
                    paint_seed=None,
                    has_stickers=None,
                    sticker_count=None,
                    inspect_link=None,
                )
            ]
        )
        snapshot_dataframe["price_source"] = commodity_price_source or "commodity_fallback"
        return attach_fetch_timestamp_columns(snapshot_dataframe)

    cheapest_listing_id, cheapest_listing, cheapest_price = min(
        priced_listings,
        key=lambda item: (item[2], item[0]),
    )
    price_source = "render_listing"
    if commodity_lowest_price is not None and commodity_lowest_price > 0:
        cheapest_price = commodity_lowest_price
        price_source = commodity_price_source or "commodity_fallback"
    elif market_level_lowest_price is not None and market_level_lowest_price > 0:
        cheapest_price = min(cheapest_price, market_level_lowest_price)
        price_source = "render_market_level"
    asset = cheapest_listing.get("asset") or {}
    asset_id = str(asset.get("id", ""))
    asset_payload = assets.get(asset_id, {})

    row = ListingRow(
        listing_id=str(cheapest_listing_id),
        asset_id=asset_id,
        page=1,
        price=cheapest_price,
        currency=str(cheapest_listing.get("currencyid", currency)),
        float_value=None,
        wear=None,
        paint_seed=None,
        has_stickers=None,
        sticker_count=None,
        inspect_link=extract_inspect_link(asset_payload, listing_id=str(cheapest_listing_id), asset_id=asset_id),
    )
    snapshot_dataframe = rows_to_dataframe([row])
    snapshot_dataframe["price_source"] = price_source
    return attach_fetch_timestamp_columns(snapshot_dataframe)


def slugify_market_hash_name(market_hash_name: str) -> str:
    market_hash_name = normalize_market_hash_name_input(market_hash_name)
    slug = re.sub(r"[^a-z0-9]+", "_", market_hash_name.lower()).strip("_")
    return slug or "steam_listings"


def classify_market_item_export_subdir(market_hash_name: str) -> Optional[str]:
    normalized_name = normalize_market_hash_name_input(market_hash_name)
    if market_item_supports_wear(normalized_name):
        return SKIN_EXPORT_SUBDIR

    if normalized_name.startswith("Sticker |"):
        return STICKER_EXPORT_SUBDIR

    if re.search(r"\bcase\b", normalized_name, re.IGNORECASE):
        return CASE_EXPORT_SUBDIR

    return None


def fetch_priceoverview_payload(
    session: requests.Session,
    market_hash_name: str,
    currency: int,
    country: str,
) -> Dict[str, Any]:
    market_hash_name = normalize_market_hash_name_input(market_hash_name)
    response = session.get(
        "https://steamcommunity.com/market/priceoverview/",
        params={
            "appid": STEAM_APP_ID,
            "market_hash_name": market_hash_name,
            "currency": currency,
            "country": country,
        },
        timeout=25,
    )
    response.raise_for_status()
    return response.json()


def fetch_listing_page_html(session: requests.Session, market_hash_name: str) -> str:
    market_hash_name = normalize_market_hash_name_input(market_hash_name)
    encoded_name = quote(market_hash_name, safe="")
    response = session.get(
        f"https://steamcommunity.com/market/listings/{STEAM_APP_ID}/{encoded_name}",
        timeout=25,
    )
    response.raise_for_status()
    return response.text


def extract_item_nameid_from_listing_html(listing_html: str) -> Optional[str]:
    match = ITEM_NAMEID_PATTERN.search(listing_html)
    if not match:
        return None
    return match.group(1)


def fetch_itemordershistogram_payload(
    session: requests.Session,
    item_nameid: str,
    currency: int,
    country: str,
    language: str,
) -> Dict[str, Any]:
    response = session.get(
        "https://steamcommunity.com/market/itemordershistogram",
        params={
            "country": country,
            "language": language,
            "currency": currency,
            "item_nameid": item_nameid,
            "two_factor": 0,
        },
        timeout=25,
    )
    response.raise_for_status()
    return response.json()


def fetch_commodity_lowest_price(
    session: requests.Session,
    market_hash_name: str,
    currency: int,
    country: str,
    language: str,
) -> tuple[Optional[float], Optional[str]]:
    listing_html = fetch_listing_page_html(session, market_hash_name)
    item_nameid = extract_item_nameid_from_listing_html(listing_html)
    if item_nameid:
        histogram_payload = fetch_itemordershistogram_payload(
            session=session,
            item_nameid=item_nameid,
            currency=currency,
            country=country,
            language=language,
        )
        if histogram_payload.get("success"):
            histogram_price = extract_lowest_histogram_price(histogram_payload)
            if histogram_price is not None and histogram_price > 0:
                return histogram_price, "itemordershistogram"

    priceoverview_payload = fetch_priceoverview_payload(
        session=session,
        market_hash_name=market_hash_name,
        currency=currency,
        country=country,
    )
    if priceoverview_payload.get("success"):
        lowest_price = parse_price_text(priceoverview_payload.get("lowest_price"))
        if lowest_price is not None and lowest_price > 0:
            return lowest_price, "priceoverview"

    return None, None


def default_fetch_output_name(market_hash_name: str) -> str:
    filename = f"{slugify_market_hash_name(market_hash_name)}.xlsx"
    subdir_name = classify_market_item_export_subdir(market_hash_name)
    if not subdir_name:
        return filename
    return str(Path(subdir_name) / filename)


def resolve_output_path(output_name: str) -> Path:
    output_path = Path(output_name)
    if output_path.parent == Path("."):
        output_path = DEFAULT_OUTPUT_DIR / output_path

    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path


def derive_output_path(input_name: str, suffix: str) -> Path:
    input_path = Path(input_name)
    extension = input_path.suffix or ".xlsx"
    derived_name = f"{input_path.stem}_{suffix}{extension}"
    return input_path.with_name(derived_name)


def get_latest_pointer_path() -> Path:
    return DEFAULT_OUTPUT_DIR / LATEST_POINTER_FILENAME


def find_newest_export_path() -> Path:
    candidate_paths = [
        path
        for path in DEFAULT_OUTPUT_DIR.rglob("*")
        if path.is_file() and path.suffix.lower() in SUPPORTED_TABLE_SUFFIXES
    ]
    if not candidate_paths:
        raise FileNotFoundError("No export files were found in exports/")
    return max(candidate_paths, key=lambda path: path.stat().st_mtime)


def read_latest_pointer() -> Optional[Path]:
    pointer_path = get_latest_pointer_path()
    if not pointer_path.exists():
        return None

    raw_target_path = pointer_path.read_text(encoding="utf-8").strip()
    if not raw_target_path:
        return None

    target_path = Path(raw_target_path)
    if target_path.exists() and target_path.is_file() and target_path.suffix.lower() in SUPPORTED_TABLE_SUFFIXES:
        return target_path

    return None


def write_latest_pointer(target_path: Path) -> Path:
    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    pointer_path = get_latest_pointer_path()
    pointer_path.write_text(str(target_path.resolve()), encoding="utf-8")
    return pointer_path


def resolve_input_path(input_name: str) -> Path:
    if input_name.lower() == "latest":
        pinned_path = read_latest_pointer()
        if pinned_path is not None:
            return pinned_path
        return find_newest_export_path()

    return Path(input_name)


def load_table(input_name: str) -> pd.DataFrame:
    input_path = resolve_input_path(input_name)
    suffix = input_path.suffix.lower()

    if suffix == ".csv":
        return pd.read_csv(input_path)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(input_path)

    raise ValueError("Input file must be a .csv, .xlsx, or .xls file")


def save_table(dataframe: pd.DataFrame, output_name: str) -> Path:
    output_path = resolve_output_path(output_name)
    suffix = output_path.suffix.lower()

    if suffix == ".csv":
        dataframe.to_csv(output_path, index=False)
        return output_path
    if suffix in {".xlsx", ".xls"}:
        dataframe.to_excel(output_path, index=False)
        try:
            format_excel_output(output_path, dataframe)
        except Exception:
            pass
        return output_path

    raise ValueError("Output file must end in .csv, .xlsx, or .xls")


def format_excel_output(output_path: Path, dataframe: pd.DataFrame) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(output_path)
    worksheet = workbook.active

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_name in enumerate(dataframe.columns, start=1):
        series = dataframe[column_name].fillna("").astype(str)
        max_data_width = max((len(value) for value in series), default=0)
        header_width = len(str(column_name))
        column_width = min(max(header_width, max_data_width) + 2, 40)
        worksheet.column_dimensions[get_column_letter(column_index)].width = column_width

    workbook.save(output_path)


def ensure_columns_exist(dataframe: pd.DataFrame, column_names: List[str]) -> None:
    missing_columns = [column for column in column_names if column not in dataframe.columns]
    if missing_columns:
        missing_text = ", ".join(missing_columns)
        raise ValueError(f"Missing required columns: {missing_text}")


def filter_dataframe(dataframe: pd.DataFrame, args: argparse.Namespace) -> pd.DataFrame:
    filtered = dataframe.copy()

    if args.min_float is not None or args.max_float is not None:
        ensure_columns_exist(filtered, ["float"])
        float_series = pd.to_numeric(filtered["float"], errors="coerce")
        if args.min_float is not None:
            filtered = filtered[float_series >= args.min_float]
            float_series = pd.to_numeric(filtered["float"], errors="coerce")
        if args.max_float is not None:
            filtered = filtered[float_series <= args.max_float]

    if args.min_price is not None or args.max_price is not None:
        ensure_columns_exist(filtered, ["price"])
        price_series = pd.to_numeric(filtered["price"], errors="coerce")
        if args.min_price is not None:
            filtered = filtered[price_series >= args.min_price]
            price_series = pd.to_numeric(filtered["price"], errors="coerce")
        if args.max_price is not None:
            filtered = filtered[price_series <= args.max_price]

    if args.wear is not None:
        ensure_columns_exist(filtered, ["wear"])
        filtered = filtered[filtered["wear"] == args.wear]

    if args.paint_seed is not None:
        ensure_columns_exist(filtered, ["paint_seed"])
        paint_seed_series = pd.to_numeric(filtered["paint_seed"], errors="coerce")
        filtered = filtered[paint_seed_series == args.paint_seed]

    if args.has_stickers:
        ensure_columns_exist(filtered, ["has_stickers"])
        sticker_series = filtered["has_stickers"].fillna(False).astype(bool)
        filtered = filtered[sticker_series]

    if args.no_stickers:
        ensure_columns_exist(filtered, ["has_stickers"])
        sticker_series = filtered["has_stickers"].fillna(False).astype(bool)
        filtered = filtered[~sticker_series]

    if args.min_sticker_count is not None or args.max_sticker_count is not None:
        ensure_columns_exist(filtered, ["sticker_count"])
        sticker_count_series = pd.to_numeric(filtered["sticker_count"], errors="coerce").fillna(0)
        if args.min_sticker_count is not None:
            filtered = filtered[sticker_count_series >= args.min_sticker_count]
            sticker_count_series = pd.to_numeric(filtered["sticker_count"], errors="coerce").fillna(0)
        if args.max_sticker_count is not None:
            filtered = filtered[sticker_count_series <= args.max_sticker_count]

    return filtered


def build_stats_lines(dataframe: pd.DataFrame, input_name: str) -> List[str]:
    lines = [f"Stats for {input_name}", f"rows: {len(dataframe)}"]

    if "price" in dataframe.columns:
        price_series = pd.to_numeric(dataframe["price"], errors="coerce").dropna()
        if not price_series.empty:
            lines.append(f"price_min: {price_series.min():.2f}")
            lines.append(f"price_max: {price_series.max():.2f}")
            lines.append(f"price_avg: {price_series.mean():.2f}")

    if "float" in dataframe.columns:
        float_series = pd.to_numeric(dataframe["float"], errors="coerce").dropna()
        if not float_series.empty:
            lines.append(f"float_min: {float_series.min():.6f}")
            lines.append(f"float_max: {float_series.max():.6f}")
            lines.append(f"float_avg: {float_series.mean():.6f}")

    if "wear" in dataframe.columns:
        wear_counts = dataframe["wear"].fillna("Unknown").value_counts()
        for wear_name, count in wear_counts.items():
            lines.append(f"wear_{wear_name}: {count}")

    if "sticker_count" in dataframe.columns:
        sticker_count_series = pd.to_numeric(dataframe["sticker_count"], errors="coerce").fillna(0)
        lines.append(f"total_stickers: {int(sticker_count_series.sum())}")

    return lines


def sort_dataframe(dataframe: pd.DataFrame, sort_by: List[str], descending: bool) -> pd.DataFrame:
    ensure_columns_exist(dataframe, sort_by)
    return dataframe.sort_values(
        by=sort_by,
        ascending=not descending,
        kind="stable",
    )


def build_show_dataframe(
    dataframe: pd.DataFrame,
    columns: Optional[List[str]] = None,
    limit: Optional[int] = None,
) -> pd.DataFrame:
    display_dataframe = dataframe.copy()

    if "has_stickers" in display_dataframe.columns and "stickers" not in display_dataframe.columns:
        display_dataframe["stickers"] = display_dataframe["has_stickers"].map(
            lambda value: "" if pd.isna(value) else ("yes" if bool(value) else "no")
        )

    if columns is None:
        columns = DEFAULT_SHOW_COLUMNS

    available_columns = [column for column in columns if column in display_dataframe.columns]
    if not available_columns:
        if display_dataframe.empty:
            display_dataframe = pd.DataFrame(columns=columns)
            available_columns = list(columns)
        else:
            raise ValueError("None of the requested display columns exist in the file")

    display_dataframe = display_dataframe[available_columns]

    if "float" in display_dataframe.columns:
        display_dataframe["float"] = pd.to_numeric(display_dataframe["float"], errors="coerce").map(
            lambda value: "" if pd.isna(value) else f"{value:.6f}"
        )
    if "price" in display_dataframe.columns:
        display_dataframe["price"] = pd.to_numeric(display_dataframe["price"], errors="coerce").map(
            lambda value: "" if pd.isna(value) else f"{value:.2f}"
        )

    display_dataframe = display_dataframe.fillna("")

    if limit is not None:
        display_dataframe = display_dataframe.head(limit)

    display_dataframe.insert(
        0,
        "#",
        range(1, len(display_dataframe) + 1),
    )

    return display_dataframe


def format_terminal_table(display_dataframe: pd.DataFrame) -> str:
    if display_dataframe.empty:
        return ""

    rendered_dataframe = display_dataframe.astype(str)
    column_widths = {
        column: max(
            len(str(column)),
            max(len(str(value)) for value in rendered_dataframe[column].tolist()),
        )
        for column in rendered_dataframe.columns
    }

    def format_cell(column: str, value: str) -> str:
        width = column_widths[column]
        if column in RIGHT_ALIGN_COLUMNS:
            return value.rjust(width)
        return value.ljust(width)

    header_line = " | ".join(
        format_cell(column, str(column))
        for column in rendered_dataframe.columns
    )
    separator_line = "-+-".join("-" * column_widths[column] for column in rendered_dataframe.columns)
    row_lines = [
        " | ".join(
            format_cell(column, str(row[column]))
            for column in rendered_dataframe.columns
        )
        for _, row in rendered_dataframe.iterrows()
    ]

    return "\n".join([header_line, separator_line, *row_lines])


def describe_listing_changes(
    previous_dataframe: Optional[pd.DataFrame],
    current_dataframe: pd.DataFrame,
) -> Optional[Dict[str, int]]:
    if previous_dataframe is None:
        return None
    if "listing_id" not in previous_dataframe.columns or "listing_id" not in current_dataframe.columns:
        return None

    previous_listing_ids = {
        str(listing_id)
        for listing_id in previous_dataframe["listing_id"].dropna().tolist()
    }
    current_listing_ids = {
        str(listing_id)
        for listing_id in current_dataframe["listing_id"].dropna().tolist()
    }

    return {
        "added": len(current_listing_ids - previous_listing_ids),
        "removed": len(previous_listing_ids - current_listing_ids),
        "unchanged": len(current_listing_ids & previous_listing_ids),
    }


def sanitize_no_wear_snapshot_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    sanitized_dataframe = dataframe.copy()
    for column_name in ("float", "wear", "paint_seed", "has_stickers", "sticker_count", "inspect_link"):
        if column_name in sanitized_dataframe.columns:
            sanitized_dataframe[column_name] = pd.NA
    return sanitized_dataframe


def append_price_snapshot_history(
    previous_dataframe: Optional[pd.DataFrame],
    current_dataframe: pd.DataFrame,
    market_hash_name: str,
) -> pd.DataFrame:
    timestamp = datetime.now().astimezone()
    snapshot_dataframe = sanitize_no_wear_snapshot_dataframe(current_dataframe)
    snapshot_dataframe["market_hash_name"] = market_hash_name
    snapshot_dataframe["snapshot_date"] = timestamp.date().isoformat()
    snapshot_dataframe["snapshot_timestamp"] = timestamp.isoformat(timespec="seconds")

    if previous_dataframe is None or previous_dataframe.empty:
        return organize_no_wear_history_dataframe(snapshot_dataframe)

    combined_dataframe = pd.concat(
        [sanitize_no_wear_snapshot_dataframe(previous_dataframe), snapshot_dataframe],
        ignore_index=True,
        sort=False,
    )
    return organize_no_wear_history_dataframe(combined_dataframe)


def organize_no_wear_history_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    organized_dataframe = dataframe.copy()

    sortable_columns = [
        column_name
        for column_name in ("market_hash_name", "snapshot_timestamp")
        if column_name in organized_dataframe.columns
    ]
    if sortable_columns:
        organized_dataframe = organized_dataframe.sort_values(
            by=sortable_columns,
            kind="stable",
        ).reset_index(drop=True)

    preferred_columns = [
        "market_hash_name",
        "snapshot_date",
        "snapshot_timestamp",
        "price",
        "price_source",
        "currency",
        "listing_id",
        "asset_id",
        "page",
        "manual_price_override",
        "manual_price_override_at",
        "float",
        "wear",
        "paint_seed",
        "has_stickers",
        "sticker_count",
        "inspect_link",
    ]
    ordered_columns = [
        column_name for column_name in preferred_columns if column_name in organized_dataframe.columns
    ]
    ordered_columns.extend(
        column_name
        for column_name in organized_dataframe.columns
        if column_name not in ordered_columns
    )
    return organized_dataframe.loc[:, ordered_columns]


def update_latest_no_wear_snapshot_price(
    output_path: Path,
    market_hash_name: str,
    new_price: float,
) -> pd.DataFrame:
    market_hash_name = normalize_market_hash_name_input(market_hash_name)
    if market_item_supports_wear(market_hash_name):
        raise ValueError("Manual latest-price updates are only supported for no-wear items")
    if new_price <= 0:
        raise ValueError("New price must be greater than zero")
    if not output_path.exists():
        raise FileNotFoundError(f"No saved export exists at {output_path}")

    dataframe = sanitize_no_wear_snapshot_dataframe(load_table(str(output_path)))
    if dataframe.empty:
        raise ValueError("Cannot update the latest price because the saved file has no rows")

    matching_row_indices = pd.Index(dataframe.index)
    if "market_hash_name" in dataframe.columns:
        normalized_market_names = dataframe["market_hash_name"].map(
            lambda value: normalize_market_hash_name_input(str(value).strip())
            if pd.notna(value)
            else ""
        )
        matching_row_indices = dataframe.index[normalized_market_names == market_hash_name]
        if matching_row_indices.empty:
            raise ValueError(
                f"Cannot update the latest price because {market_hash_name} has no saved rows in {output_path}"
            )
    latest_row_index = matching_row_indices[-1]

    if "price" not in dataframe.columns:
        dataframe["price"] = pd.NA
    dataframe.loc[latest_row_index, "price"] = float(new_price)

    if "price_source" not in dataframe.columns:
        dataframe["price_source"] = pd.NA
    dataframe.loc[latest_row_index, "price_source"] = "manual_override"

    if "manual_price_override" not in dataframe.columns:
        dataframe["manual_price_override"] = pd.NA
    dataframe.loc[latest_row_index, "manual_price_override"] = True

    if "manual_price_override_at" not in dataframe.columns:
        dataframe["manual_price_override_at"] = pd.NA
    dataframe.loc[latest_row_index, "manual_price_override_at"] = datetime.now().astimezone().isoformat(timespec="seconds")

    organized_dataframe = organize_no_wear_history_dataframe(dataframe)
    save_table(organized_dataframe, str(output_path))
    return organized_dataframe


def create_requests_session() -> requests.Session:
    session = requests.Session()
    # Steam access is more reliable when we bypass any broken local proxy env vars.
    # This tool is intended to talk directly to Steam, not through a custom proxy.
    session.trust_env = False
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        }
    )
    return session


def close_requests_session(session: requests.Session) -> None:
    try:
        session.close()
    except Exception:
        pass


def should_refresh_steam_session(pages_fetched: int) -> bool:
    return (
        pages_fetched > 0
        and pages_fetched % STEAM_SESSION_REFRESH_PAGE_INTERVAL == 0
    )


def recover_steam_render_page(
    *,
    session: requests.Session,
    market_hash_name: str,
    start: int,
    currency: int,
    country: str,
    language: str,
    max_retries: int,
    steam_page_delay: float,
) -> tuple[Dict[str, Any], requests.Session, float]:
    effective_page_delay = max(steam_page_delay, STEAM_RECOVERY_PAGE_DELAY_FLOOR)
    current_session = session

    for recovery_index, base_wait_seconds in enumerate(STEAM_RECOVERY_WAIT_STEPS, start=1):
        wait_seconds = max(base_wait_seconds, effective_page_delay)
        print(
            f"Steam kept rejecting the page starting at {start}. "
            f"Cooling down for {wait_seconds:.1f}s and refreshing the session "
            f"(recovery {recovery_index}/{len(STEAM_RECOVERY_WAIT_STEPS)})..."
        )
        time.sleep(wait_seconds)
        current_session = create_requests_session()
        try:
            payload = steam_render_page(
                session=current_session,
                market_hash_name=market_hash_name,
                start=start,
                currency=currency,
                country=country,
                language=language,
                max_retries=max_retries,
            )
            return payload, current_session, effective_page_delay
        except requests.HTTPError as exc:
            response = getattr(exc, "response", None)
            if getattr(response, "status_code", None) not in RETRIABLE_STATUS_CODES:
                raise
            close_requests_session(current_session)

    raise requests.HTTPError(
        f"Steam kept rate-limiting page start={start} even after extended recovery",
        response=response,
    )


def dataframe_matches_inline_query(dataframe: pd.DataFrame, args: argparse.Namespace) -> pd.DataFrame:
    if dataframe.empty:
        empty_dataframe = dataframe.copy()
        if empty_dataframe.columns.empty and getattr(args, "sort_by", None):
            empty_dataframe = pd.DataFrame(columns=list(args.sort_by))
        return empty_dataframe

    filtered_dataframe = filter_dataframe(dataframe, args)
    if getattr(args, "sort_by", None):
        filtered_dataframe = sort_dataframe(filtered_dataframe, args.sort_by, args.descending)
    return filtered_dataframe


def has_inline_fetch_query(args: argparse.Namespace) -> bool:
    return any(
        getattr(args, attribute_name, None)
        for attribute_name in (
            "min_float",
            "max_float",
            "min_price",
            "max_price",
            "wear",
            "paint_seed",
            "has_stickers",
            "no_stickers",
            "min_sticker_count",
            "max_sticker_count",
            "sort_by",
            "show",
        )
    )


def print_matching_rows(
    label: str,
    filtered_dataframe: pd.DataFrame,
    limit: Optional[int],
    columns: Optional[List[str]],
) -> None:
    display_dataframe = build_show_dataframe(
        filtered_dataframe,
        columns=columns,
        limit=limit,
    )
    print(f"Showing {len(display_dataframe)} of {len(filtered_dataframe)} matching rows from {label}")
    if display_dataframe.empty:
        return

    print(format_terminal_table(display_dataframe))


def print_fetch_inline_summary(
    market_hash_name: str,
    output_path: Path,
    dataframe: pd.DataFrame,
    args: argparse.Namespace,
) -> None:
    if not has_inline_fetch_query(args):
        return

    filtered_dataframe = dataframe_matches_inline_query(dataframe, args)
    print(
        f"Inline query matched {len(filtered_dataframe)} rows for {market_hash_name} "
        f"from {output_path}"
    )

    if args.show:
        print_matching_rows(
            label=f"{market_hash_name} ({output_path.name})",
            filtered_dataframe=filtered_dataframe,
            limit=args.limit,
            columns=args.columns,
        )


def build_fetch_result_summary(result: FetchResult) -> str:
    if result.summary_override:
        return result.summary_override
    if result.change_summary is None:
        return f"Exported {len(result.dataframe)} listings to {result.output_path}"

    return (
        f"Synced {len(result.dataframe)} current listings to {result.output_path} "
        f"(added {result.change_summary['added']}, removed {result.change_summary['removed']}, "
        f"unchanged {result.change_summary['unchanged']})"
    )


def fetch_market_dataframe(args: argparse.Namespace, market_hash_name: str) -> pd.DataFrame:
    market_hash_name = normalize_market_hash_name_input(market_hash_name)
    session = create_requests_session()
    try:
        if not market_item_supports_wear(market_hash_name):
            return build_lowest_listing_snapshot_dataframe(
                session=session,
                market_hash_name=market_hash_name,
                currency=args.currency,
                country=args.country,
                language=args.language,
                steam_max_retries=args.steam_max_retries,
            )

        rows = list(
            iter_listings(
                session=session,
                market_hash_name=market_hash_name,
                currency=args.currency,
                country=args.country,
                language=args.language,
                steam_page_delay=args.steam_page_delay,
                steam_max_retries=args.steam_max_retries,
            )
        )
        return attach_fetch_timestamp_columns(rows_to_dataframe(rows))
    finally:
        close_requests_session(session)


def sync_market_dataframe(
    dataframe: pd.DataFrame,
    market_hash_name: str,
    output_name: Optional[str] = None,
    update_latest: bool = True,
) -> FetchResult:
    market_hash_name = normalize_market_hash_name_input(market_hash_name)
    resolved_output_name = output_name or default_fetch_output_name(market_hash_name)
    if output_name is None:
        output_path = resolve_output_path(str(DEFAULT_OUTPUT_DIR / Path(resolved_output_name)))
    else:
        output_path = resolve_output_path(resolved_output_name)
    previous_dataframe = load_table(str(output_path)) if output_path.exists() else None
    summary_override: Optional[str] = None

    if not market_item_supports_wear(market_hash_name):
        if dataframe.empty:
            if previous_dataframe is not None:
                final_dataframe = sanitize_no_wear_snapshot_dataframe(previous_dataframe)
                summary_override = (
                    f"Steam did not return a current active listing price for {market_hash_name}. "
                    f"Kept {len(final_dataframe)} historical snapshot row(s) in {output_path}"
                )
            else:
                final_dataframe = dataframe.copy()
                summary_override = (
                    f"Steam did not return a current active listing price for {market_hash_name}, "
                    f"so no snapshot row was added."
                )
        else:
            final_dataframe = append_price_snapshot_history(
                previous_dataframe=previous_dataframe,
                current_dataframe=dataframe,
                market_hash_name=market_hash_name,
            )
            appended_rows = len(dataframe)
            summary_override = (
                f"Appended {appended_rows} lowest-price snapshot row(s) for {market_hash_name} "
                f"into {output_path} (history rows: {len(final_dataframe)})"
            )
    else:
        final_dataframe = dataframe

    save_table(final_dataframe, str(output_path))
    if update_latest:
        write_latest_pointer(output_path)

    return FetchResult(
        market_hash_name=market_hash_name,
        output_path=output_path,
        dataframe=final_dataframe,
        change_summary=describe_listing_changes(previous_dataframe, final_dataframe)
        if market_item_supports_wear(market_hash_name)
        else None,
        summary_override=summary_override,
    )


def add_fetch_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "market_hash_name",
        help='Exact Steam market hash name, e.g. "AK-47 | Redline (Field-Tested)"',
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output Excel filename. If omitted, the tool derives one from the market name inside exports/",
    )
    parser.add_argument("--currency", type=int, default=1,
                        help="Steam currency ID (default: 1 for USD)")
    parser.add_argument("--country", default="US",
                        help="Steam country code (default: US)")
    parser.add_argument("--language", default="english",
                        help="Steam language (default: english)")
    parser.add_argument(
        "--steam-page-delay",
        type=float,
        default=DEFAULT_STEAM_PAGE_DELAY,
        help="Seconds to wait between Steam listing page requests (default: 0.0)",
    )
    parser.add_argument(
        "--steam-max-retries",
        type=int,
        default=DEFAULT_STEAM_RETRIES,
        help="How many times to retry a Steam page after HTTP 429 (default: 5)",
    )
    parser.add_argument("--min-float", type=float, default=None, help="After fetching, keep rows with float >= this value")
    parser.add_argument("--max-float", type=float, default=None, help="After fetching, keep rows with float <= this value")
    parser.add_argument("--min-price", type=float, default=None, help="After fetching, keep rows with price >= this value")
    parser.add_argument("--max-price", type=float, default=None, help="After fetching, keep rows with price <= this value")
    parser.add_argument("--wear", default=None, help="After fetching, keep only rows with this wear value")
    parser.add_argument("--paint-seed", type=int, default=None, help="After fetching, keep only rows with this paint seed")

    sticker_group = parser.add_mutually_exclusive_group()
    sticker_group.add_argument(
        "--has-stickers",
        action="store_true",
        help="After fetching, keep only rows that have stickers",
    )
    sticker_group.add_argument(
        "--no-stickers",
        action="store_true",
        help="After fetching, keep only rows that do not have stickers",
    )

    parser.add_argument(
        "--min-sticker-count",
        type=int,
        default=None,
        help="After fetching, keep rows with sticker_count >= this value",
    )
    parser.add_argument(
        "--max-sticker-count",
        type=int,
        default=None,
        help="After fetching, keep rows with sticker_count <= this value",
    )
    parser.add_argument(
        "--sort-by",
        nargs="+",
        default=None,
        help="After fetching, sort matching rows by these columns",
    )
    parser.add_argument(
        "--descending",
        action="store_true",
        help="Use descending order for inline fetch sorting",
    )
    parser.add_argument(
        "--show",
        action="store_true",
        help="After fetching, print the matching rows directly in the terminal",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=25,
        help="Maximum number of inline fetch rows to show in the terminal (default: 25)",
    )
    parser.add_argument(
        "--columns",
        nargs="+",
        default=None,
        help="Column names to show in inline fetch terminal output",
    )


def add_fetch_many_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "market_hash_names",
        nargs="*",
        help='One or more exact Steam market names, e.g. "AK-47 | Redline (Field-Tested)"',
    )
    parser.add_argument(
        "--items-file",
        default=None,
        help="Optional text file with one exact Steam market name per line",
    )
    parser.add_argument("--currency", type=int, default=1,
                        help="Steam currency ID (default: 1 for USD)")
    parser.add_argument("--country", default="US",
                        help="Steam country code (default: US)")
    parser.add_argument("--language", default="english",
                        help="Steam language (default: english)")
    parser.add_argument(
        "--steam-page-delay",
        type=float,
        default=DEFAULT_STEAM_PAGE_DELAY,
        help="Seconds to wait between Steam listing page requests inside each worker (default: 0.0)",
    )
    parser.add_argument(
        "--steam-max-retries",
        type=int,
        default=DEFAULT_STEAM_RETRIES,
        help="How many times to retry a Steam page after HTTP 429 (default: 5)",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=DEFAULT_FETCH_MANY_WORKERS,
        help="How many market items to fetch in parallel (default: 3)",
    )
    parser.add_argument("--min-float", type=float, default=None, help="After fetching, keep rows with float >= this value")
    parser.add_argument("--max-float", type=float, default=None, help="After fetching, keep rows with float <= this value")
    parser.add_argument("--min-price", type=float, default=None, help="After fetching, keep rows with price >= this value")
    parser.add_argument("--max-price", type=float, default=None, help="After fetching, keep rows with price <= this value")
    parser.add_argument("--wear", default=None, help="After fetching, keep only rows with this wear value")
    parser.add_argument("--paint-seed", type=int, default=None, help="After fetching, keep only rows with this paint seed")

    sticker_group = parser.add_mutually_exclusive_group()
    sticker_group.add_argument(
        "--has-stickers",
        action="store_true",
        help="After fetching, keep only rows that have stickers",
    )
    sticker_group.add_argument(
        "--no-stickers",
        action="store_true",
        help="After fetching, keep only rows that do not have stickers",
    )

    parser.add_argument(
        "--min-sticker-count",
        type=int,
        default=None,
        help="After fetching, keep rows with sticker_count >= this value",
    )
    parser.add_argument(
        "--max-sticker-count",
        type=int,
        default=None,
        help="After fetching, keep rows with sticker_count <= this value",
    )
    parser.add_argument(
        "--sort-by",
        nargs="+",
        default=None,
        help="After fetching, sort matching rows by these columns",
    )
    parser.add_argument(
        "--descending",
        action="store_true",
        help="Use descending order for inline fetch-many sorting",
    )
    parser.add_argument(
        "--show",
        action="store_true",
        help="After fetching each item, print the matching rows directly in the terminal",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=25,
        help="Maximum number of inline fetch-many rows to show per item in the terminal (default: 25)",
    )
    parser.add_argument(
        "--columns",
        nargs="+",
        default=None,
        help="Column names to show in inline fetch-many terminal output",
    )


def add_input_path_argument(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "input_path",
        help="Path to an existing .xlsx, .xls, or .csv file created by this tool, or use latest",
    )


def add_sort_arguments(parser: argparse.ArgumentParser) -> None:
    add_input_path_argument(parser)
    parser.add_argument(
        "--by",
        nargs="+",
        required=True,
        help="One or more column names to sort by, e.g. --by float price",
    )
    parser.add_argument(
        "--descending",
        action="store_true",
        help="Sort in descending order instead of ascending order",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output file path. Default is based on the input file name.",
    )


def add_filter_arguments(parser: argparse.ArgumentParser) -> None:
    add_input_path_argument(parser)
    parser.add_argument("--min-float", type=float, default=None, help="Keep rows with float >= this value")
    parser.add_argument("--max-float", type=float, default=None, help="Keep rows with float <= this value")
    parser.add_argument("--min-price", type=float, default=None, help="Keep rows with price >= this value")
    parser.add_argument("--max-price", type=float, default=None, help="Keep rows with price <= this value")
    parser.add_argument("--wear", default=None, help="Keep only rows with this wear value")
    parser.add_argument("--paint-seed", type=int, default=None, help="Keep only rows with this paint seed")

    sticker_group = parser.add_mutually_exclusive_group()
    sticker_group.add_argument(
        "--has-stickers",
        action="store_true",
        help="Keep only rows that have stickers",
    )
    sticker_group.add_argument(
        "--no-stickers",
        action="store_true",
        help="Keep only rows that do not have stickers",
    )

    parser.add_argument(
        "--min-sticker-count",
        type=int,
        default=None,
        help="Keep rows with sticker_count >= this value",
    )
    parser.add_argument(
        "--max-sticker-count",
        type=int,
        default=None,
        help="Keep rows with sticker_count <= this value",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output file path. Default is based on the input file name.",
    )


def add_show_arguments(parser: argparse.ArgumentParser) -> None:
    add_input_path_argument(parser)
    parser.add_argument("--min-float", type=float, default=None, help="Show rows with float >= this value")
    parser.add_argument("--max-float", type=float, default=None, help="Show rows with float <= this value")
    parser.add_argument("--min-price", type=float, default=None, help="Show rows with price >= this value")
    parser.add_argument("--max-price", type=float, default=None, help="Show rows with price <= this value")
    parser.add_argument("--wear", default=None, help="Show only rows with this wear value")
    parser.add_argument("--paint-seed", type=int, default=None, help="Show only rows with this paint seed")

    sticker_group = parser.add_mutually_exclusive_group()
    sticker_group.add_argument(
        "--has-stickers",
        action="store_true",
        help="Show only rows that have stickers",
    )
    sticker_group.add_argument(
        "--no-stickers",
        action="store_true",
        help="Show only rows that do not have stickers",
    )

    parser.add_argument(
        "--min-sticker-count",
        type=int,
        default=None,
        help="Show rows with sticker_count >= this value",
    )
    parser.add_argument(
        "--max-sticker-count",
        type=int,
        default=None,
        help="Show rows with sticker_count <= this value",
    )
    parser.add_argument(
        "--sort-by",
        nargs="+",
        default=None,
        help="One or more column names to sort by before showing rows",
    )
    parser.add_argument(
        "--descending",
        action="store_true",
        help="Show rows in descending sort order",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=25,
        help="Maximum number of rows to show in the terminal (default: 25)",
    )
    parser.add_argument(
        "--columns",
        nargs="+",
        default=None,
        help="Column names to show in the terminal output",
    )


def add_stats_arguments(parser: argparse.ArgumentParser) -> None:
    add_input_path_argument(parser)


def add_use_arguments(parser: argparse.ArgumentParser) -> None:
    add_input_path_argument(parser)


def build_legacy_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Scrape Steam Community Market listings for a CS2 item and export to Excel."
    )
    add_fetch_arguments(parser)
    parser.set_defaults(command="fetch")
    return parser


def build_cli_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Steam Market Listing to Excel CLI. "
            "Legacy fetch style still works if you pass the market name directly."
        )
    )
    subparsers = parser.add_subparsers(dest="command")

    fetch_parser = subparsers.add_parser(
        "fetch",
        help="Scrape Steam Community Market listings and export to Excel.",
        description="Scrape Steam Community Market listings for a CS2 item and export to Excel.",
    )
    add_fetch_arguments(fetch_parser)

    fetch_many_parser = subparsers.add_parser(
        "fetch-many",
        help="Fetch multiple Steam Community Market items in parallel.",
        description="Fetch multiple Steam Community Market items in parallel.",
    )
    add_fetch_many_arguments(fetch_many_parser)

    sort_parser = subparsers.add_parser(
        "sort",
        help="Sort an existing export file and write a new file.",
        description="Sort an existing export file and write a new file.",
    )
    add_sort_arguments(sort_parser)

    filter_parser = subparsers.add_parser(
        "filter",
        help="Filter an existing export file and write a new file.",
        description="Filter an existing export file and write a new file.",
    )
    add_filter_arguments(filter_parser)

    show_parser = subparsers.add_parser(
        "show",
        help="Show matching rows from an existing export file in the terminal.",
        description="Show matching rows from an existing export file in the terminal.",
    )
    add_show_arguments(show_parser)

    stats_parser = subparsers.add_parser(
        "stats",
        help="Print a quick summary of an existing export file.",
        description="Print a quick summary of an existing export file.",
    )
    add_stats_arguments(stats_parser)

    use_parser = subparsers.add_parser(
        "use",
        help="Set which export file the latest shortcut should point to.",
        description="Set which export file the latest shortcut should point to.",
    )
    add_use_arguments(use_parser)

    return parser


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    if argv is None:
        argv = sys.argv[1:]

    if argv and (argv[0] in CLI_COMMANDS or argv[0] in {"-h", "--help"}):
        return build_cli_parser().parse_args(argv)

    return build_legacy_parser().parse_args(argv)


def run_fetch(args: argparse.Namespace) -> None:
    args.market_hash_name = normalize_market_hash_name_input(args.market_hash_name)
    dataframe = fetch_market_dataframe(args, args.market_hash_name)
    result = sync_market_dataframe(
        dataframe=dataframe,
        market_hash_name=args.market_hash_name,
        output_name=args.output,
        update_latest=True,
    )
    print(build_fetch_result_summary(result))
    print_fetch_inline_summary(
        market_hash_name=args.market_hash_name,
        output_path=result.output_path,
        dataframe=result.dataframe,
        args=args,
    )


def collect_market_hash_names(args: argparse.Namespace) -> List[str]:
    market_hash_names = [
        normalize_market_hash_name_input(market_hash_name)
        for market_hash_name in list(args.market_hash_names or [])
    ]

    if args.items_file:
        items_file_path = Path(args.items_file)
        file_market_hash_names = [
            normalize_market_hash_name_input(line.strip())
            for line in items_file_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        market_hash_names.extend(file_market_hash_names)

    deduplicated_market_hash_names: List[str] = []
    seen_market_hash_names = set()
    for market_hash_name in market_hash_names:
        if market_hash_name not in seen_market_hash_names:
            deduplicated_market_hash_names.append(market_hash_name)
            seen_market_hash_names.add(market_hash_name)

    if not deduplicated_market_hash_names:
        raise ValueError("fetch-many needs at least one market name or an --items-file")

    return deduplicated_market_hash_names


def run_fetch_many(args: argparse.Namespace) -> None:
    market_hash_names = collect_market_hash_names(args)
    max_workers = max(1, min(args.workers, len(market_hash_names)))
    print(f"Fetching {len(market_hash_names)} items with {max_workers} worker(s)...")

    results_by_market_name: Dict[str, FetchResult] = {}
    errors_by_market_name: Dict[str, Exception] = {}

    def fetch_one_item(market_hash_name: str) -> FetchResult:
        dataframe = fetch_market_dataframe(args, market_hash_name)
        return sync_market_dataframe(
            dataframe=dataframe,
            market_hash_name=market_hash_name,
            output_name=None,
            update_latest=False,
        )

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_market_name = {
            executor.submit(fetch_one_item, market_hash_name): market_hash_name
            for market_hash_name in market_hash_names
        }

        for future in as_completed(future_to_market_name):
            market_hash_name = future_to_market_name[future]
            try:
                result = future.result()
                results_by_market_name[market_hash_name] = result
                print(build_fetch_result_summary(result))
                print_fetch_inline_summary(
                    market_hash_name=market_hash_name,
                    output_path=result.output_path,
                    dataframe=result.dataframe,
                    args=args,
                )
            except Exception as exc:  # pragma: no cover - exercised via behavior, not exact type branching
                errors_by_market_name[market_hash_name] = exc
                print(f"Failed to fetch {market_hash_name}: {exc}")

    for market_hash_name in reversed(market_hash_names):
        if market_hash_name in results_by_market_name:
            write_latest_pointer(results_by_market_name[market_hash_name].output_path)
            break

    if errors_by_market_name:
        failed_count = len(errors_by_market_name)
        success_count = len(results_by_market_name)
        raise RuntimeError(
            f"fetch-many finished with {success_count} success(es) and {failed_count} failure(s)"
        )


def run_sort(args: argparse.Namespace) -> None:
    resolved_input_path = resolve_input_path(args.input_path)
    dataframe = load_table(str(resolved_input_path))
    sorted_dataframe = sort_dataframe(dataframe, args.by, args.descending)
    output_name = args.output or str(derive_output_path(str(resolved_input_path), "sorted"))
    output_path = save_table(sorted_dataframe, output_name)
    write_latest_pointer(output_path)

    print(f"Sorted {len(sorted_dataframe)} rows into {output_path}")


def run_filter(args: argparse.Namespace) -> None:
    resolved_input_path = resolve_input_path(args.input_path)
    dataframe = load_table(str(resolved_input_path))
    filtered_dataframe = filter_dataframe(dataframe, args)

    output_name = args.output or str(derive_output_path(str(resolved_input_path), "filtered"))
    output_path = save_table(filtered_dataframe, output_name)
    write_latest_pointer(output_path)

    print(f"Filtered {len(filtered_dataframe)} rows into {output_path}")


def run_stats(args: argparse.Namespace) -> None:
    resolved_input_path = resolve_input_path(args.input_path)
    dataframe = load_table(str(resolved_input_path))
    for line in build_stats_lines(dataframe, str(resolved_input_path)):
        print(line)


def run_show(args: argparse.Namespace) -> None:
    resolved_input_path = resolve_input_path(args.input_path)
    dataframe = load_table(str(resolved_input_path))
    filtered_dataframe = filter_dataframe(dataframe, args)

    if args.sort_by:
        filtered_dataframe = sort_dataframe(filtered_dataframe, args.sort_by, args.descending)

    print_matching_rows(
        label=str(resolved_input_path),
        filtered_dataframe=filtered_dataframe,
        limit=args.limit,
        columns=args.columns,
    )


def run_use(args: argparse.Namespace) -> None:
    resolved_input_path = resolve_input_path(args.input_path)
    if not resolved_input_path.exists() or not resolved_input_path.is_file():
        raise FileNotFoundError(f"Input file was not found: {resolved_input_path}")
    if resolved_input_path.suffix.lower() not in SUPPORTED_TABLE_SUFFIXES:
        raise ValueError("Input file must be a .csv, .xlsx, or .xls file")

    write_latest_pointer(resolved_input_path)
    print(f"latest now points to {resolved_input_path.resolve()}")


def main(argv: Optional[List[str]] = None) -> None:
    args = parse_args(argv)
    if args.command == "fetch":
        run_fetch(args)
    elif args.command == "fetch-many":
        run_fetch_many(args)
    elif args.command == "sort":
        run_sort(args)
    elif args.command == "filter":
        run_filter(args)
    elif args.command == "show":
        run_show(args)
    elif args.command == "stats":
        run_stats(args)
    elif args.command == "use":
        run_use(args)


if __name__ == "__main__":
    main()
